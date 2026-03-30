import streamlit as st
import pandas as pd
from io import BytesIO
from collections import Counter
from openpyxl import load_workbook
import os
import re
import difflib

st.set_page_config(page_title="타운보드 센터 배정용", layout="wide")
st.title("📍타운보드 센터 배정용")
st.info("""
**작동 방식**
- 가동리스트는 자동으로 불러옵니다.
- 광고게첨리스트 파일 → **O열(센터명) + S열(아파트명)** 자동 입력
- 송출요청서 파일명과 광고명이 완전히 일치하지 않아도 **유사도 50% 이상**이면 자동 매칭합니다.
- 단, 차수(1차·2차 등)가 서로 다른 경우는 매칭하지 않습니다.
""")

A_FILE = "타운보드_가동리스트_패키지상품__260323.xlsx"

@st.cache_data
def load_apt_map():
    if not os.path.exists(A_FILE):
        return None, f"❌ A파일({A_FILE})을 찾을 수 없습니다."
    try:
        with open(A_FILE, "rb") as f:
            file_bytes = BytesIO(f.read())
        wb = load_workbook(file_bytes, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        result = {}
        header_found = False
        col_apt, col_g, col_center = 3, 7, 13

        for row in ws.iter_rows(values_only=True):
            if not header_found:
                row_vals = [str(v) if v is not None else '' for v in row]
                if '아파트명' in row_vals:
                    col_apt    = row_vals.index('아파트명')
                    col_g      = row_vals.index('지역3(법정)') if '지역3(법정)' in row_vals else 7
                    col_center = row_vals.index('센터명') if '센터명' in row_vals else 13
                    header_found = True
                continue

            aname  = str(row[col_apt]).strip()    if row[col_apt]    is not None else ''
            g_val  = str(row[col_g]).strip()      if row[col_g]      is not None else ''
            center = str(row[col_center]).strip() if row[col_center] is not None else ''

            if not aname or aname in ('None', '아파트명') or aname.startswith('합'):
                continue
            if center in ('Y', 'None'):
                center = ''

            if aname not in result:
                result[aname] = (g_val, center)
            elif not result[aname][1] and center:
                result[aname] = (g_val, center)

        wb.close()
        return result, None
    except Exception as e:
        return None, f"❌ A파일 로드 오류: {e}"


apt_map, err = load_apt_map()
if err:
    st.error(err)
    st.stop()

센터있음 = sum(1 for v in apt_map.values() if v[1])
st.success(f"✅ 가동리스트 로드 완료 → 아파트 {len(apt_map):,}개 (센터명 매핑: {센터있음:,}개)")

SKIP = {'단지명', '매체그룹명', 'nan', '', '합  계', '합계', '아파트', '구분',
        '단지수', '소재명', '광고주', 'MGID', '유형별', '선택', '아파트명', 'None'}

SIMILARITY_THRESHOLD = 0.5  # 유사도 임계값

# ── 유사도 매칭 함수 ──────────────────────────────────────────
def _normalize(s):
    """매칭용 정규화: 날짜·괄호·공백·특수문자 제거"""
    s = s.strip()
    s = re.sub(r'[()（）\[\]]', '_', s)          # 괄호 → 언더스코어
    s = re.sub(r'_?\d{6}_?', '', s)              # 6자리 날짜 (250625)
    s = re.sub(r'_?\d{4}[_년]\d{1,2}[월_]?', '', s)  # 연도_월 (2026_02월, 26년2월)
    s = re.sub(r'[_\s]+', '', s)                 # 공백·언더스코어 제거
    return s.lower()

def _extract_ordinal(s):
    """차수 숫자 추출: '2차' → 2, 없으면 None"""
    m = re.search(r'(\d+)차', s)
    return int(m.group(1)) if m else None

def _similarity(ad_name, key):
    """
    광고명 ↔ 파일키 유사도 계산 (0.0 ~ 1.0)
    - 차수가 둘 다 있고 다르면 → 0.0 (강제 불일치)
    - 파일키에 차수 있는데 광고명엔 없으면 → 0.0 (다른 버전으로 간주)
    - 그 외 → 정규화 후 문자열 유사도
    """
    ord_a = _extract_ordinal(ad_name)
    ord_k = _extract_ordinal(key)

    if ord_k is not None and ord_a is None:
        return 0.0  # 파일은 2차인데 광고명엔 차수 없음 → 다른 광고
    if ord_a is not None and ord_k is not None and ord_a != ord_k:
        return 0.0  # 차수 충돌

    na, nk = _normalize(ad_name), _normalize(key)
    return difflib.SequenceMatcher(None, na, nk).ratio()

def fuzzy_match(ad_name, keys):
    """
    광고명과 파일키 목록을 비교해 가장 유사한 키 반환.
    1순위: 정확 일치
    2순위: 유사도 ≥ 0.5 중 가장 높은 것
    매칭 없으면 None
    """
    ad_clean = ad_name.strip()

    # 1순위: 정확 일치
    for k in keys:
        if ad_clean == k:
            return k, 1.0

    # 2순위: 유사도 기반
    scores = [(k, _similarity(ad_clean, k)) for k in keys]
    scores = [(k, s) for k, s in scores if s >= SIMILARITY_THRESHOLD]
    if scores:
        best_key, best_score = max(scores, key=lambda x: x[1])
        return best_key, best_score

    return None, 0.0


def find_header_row(ws):
    """헤더 행 탐색 → (row 번호 1-indexed, 광고명 열 0-indexed)"""
    for i, row in enumerate(ws.iter_rows(max_row=20, values_only=True), start=1):
        row_vals = [''.join(str(v).split()) if v is not None else '' for v in row]
        for j, v in enumerate(row_vals):
            if '소재명' in v:
                return i, j
    return 1, 5

def find_apts_auto(uploaded_file, apt_map):
    """광고파일에서 A파일 매칭 아파트를 가장 많이 찾을 수 있는 시트·열 자동 탐색"""
    try:
        xl = pd.ExcelFile(uploaded_file)
    except Exception:
        return [], None, None
    best_count, best_apts, best_sheet, best_col = 0, [], None, None
    for sheet in xl.sheet_names:
        try:
            uploaded_file.seek(0)
            df = pd.read_excel(uploaded_file, sheet_name=sheet, header=None)
            for col_idx in range(min(12, df.shape[1])):
                col_vals = df.iloc[:, col_idx].dropna().astype(str).str.strip()
                matches = [v for v in col_vals
                           if v not in SKIP and not v.startswith('합') and v in apt_map]
                unique = list(dict.fromkeys(matches))
                if len(unique) > best_count:
                    best_count = len(unique)
                    best_apts = unique
                    best_sheet = sheet
                    best_col = col_idx
        except Exception:
            continue
    try:
        uploaded_file.seek(0)
    except Exception:
        pass
    return best_apts, best_sheet, best_col

def ad_key(fname):
    return fname.replace('.xlsx', '').replace('송출요청서_', '').strip()

# 대구 패키지 센터 우선순위
DAEGU_PKG_CENTER_PRIORITY = ['칠성', '대남', '월배']

def is_daegu_package(ws, excel_row):
    """M열=대구 AND P열(아파트지정)에 '패키지' 또는 'Half' 포함 여부"""
    m_val = ws.cell(row=excel_row, column=13).value  # M열: 지역
    p_val = ws.cell(row=excel_row, column=16).value  # P열: 아파트지정
    m_str = str(m_val).strip() if m_val else ''
    p_str = str(p_val).strip() if p_val else ''
    return '대구' in m_str and ('패키지' in p_str or 'Half' in p_str)

def pick_daegu_pkg_apt(apt_map, used_centers):
    """칠성→대남→월배 순으로 미사용 센터에서 아파트 반환"""
    for center in DAEGU_PKG_CENTER_PRIORITY:
        if center not in used_centers:
            center_apts = [a for a, (g, c) in apt_map.items() if c == center]
            if center_apts:
                return center_apts[0], center
    # 모두 소진 시 칠성 재사용
    center_apts = [a for a, (g, c) in apt_map.items() if c == '칠성']
    return (center_apts[0], '칠성') if center_apts else (None, None)

def process_b_file(b_file, ad_file_apts, apt_map, apt_freq):
    """게첨리스트 B파일 처리 → (결과 wb, results 리스트)"""
    b_file.seek(0)
    wb = load_workbook(b_file)
    ws = wb.active

    header_row, ad_col_idx = find_header_row(ws)
    ad_col = ad_col_idx + 1

    COL_CENTER = 15   # O열
    COL_APT    = 19   # S열

    used_g_per_ad = {}
    used_centers_per_ad = {}  # 대구 패키지용
    results = []

    for excel_row in range(header_row + 1, ws.max_row + 1):
        cell_f = ws.cell(row=excel_row, column=ad_col)
        ad_name = str(cell_f.value).strip() if cell_f.value else ''
        if not ad_name or ''.join(ad_name.split()) in ('', '소재명'):
            continue

        # ── 대구 패키지 특별 처리 ──────────────────────────────
        if is_daegu_package(ws, excel_row):
            if ad_name not in used_centers_per_ad:
                used_centers_per_ad[ad_name] = set()
            chosen, center = pick_daegu_pkg_apt(apt_map, used_centers_per_ad[ad_name])
            if chosen:
                used_centers_per_ad[ad_name].add(center)
                ws.cell(row=excel_row, column=COL_CENTER).value = center
                ws.cell(row=excel_row, column=COL_APT).value = chosen
                results.append({'행': excel_row, '광고명': ad_name,
                                '매칭 파일': '(대구 패키지 규칙)',
                                '센터명': center, '아파트명': chosen,
                                '유사도': '—', '비고': f'대구패키지 우선배정 ({center})'})
            else:
                results.append({'행': excel_row, '광고명': ad_name,
                                '매칭 파일': '—', '센터명': '—', '아파트명': '—',
                                '유사도': '—', '비고': '대구패키지 배정 실패'})
            continue
        # ──────────────────────────────────────────────────────

        matched_key, score = fuzzy_match(ad_name, list(ad_file_apts.keys()))

        if not matched_key:
            results.append({'행': excel_row, '광고명': ad_name,
                            '센터명': '—', '아파트명': '—',
                            '유사도': '—', '비고': '송출요청서 없음'})
            continue

        candidates = ad_file_apts.get(matched_key, [])
        score_str = '정확일치' if score == 1.0 else f'{score:.0%}'

        if not candidates:
            results.append({'행': excel_row, '광고명': ad_name,
                            '센터명': '—', '아파트명': '—',
                            '유사도': score_str, '비고': '패키지 상품 (아파트 미지정)'})
            continue

        if matched_key not in used_g_per_ad:
            used_g_per_ad[matched_key] = set()

        sorted_cands = sorted(candidates, key=lambda a: apt_freq.get(a, 0), reverse=True)
        chosen = None
        for apt in sorted_cands:
            g_val, _ = apt_map[apt]
            if g_val not in used_g_per_ad[matched_key]:
                chosen = apt
                used_g_per_ad[matched_key].add(g_val)
                break
        if not chosen:
            chosen = sorted_cands[0]

        g_val, center = apt_map[chosen]
        ws.cell(row=excel_row, column=COL_CENTER).value = center if center else ''
        ws.cell(row=excel_row, column=COL_APT).value = chosen
        results.append({'행': excel_row, '광고명': ad_name,
                        '매칭 파일': matched_key,
                        '센터명': center if center else '(센터없음)',
                        '아파트명': chosen,
                        '유사도': score_str,
                        '비고': f"공통 {apt_freq.get(chosen, 1)}개 광고"})

    return wb, results


# ── UI ──
st.divider()
col1, col2 = st.columns(2)
with col1:
    b_files = st.file_uploader(
        "📋 광고게첨리스트 파일 (여러 개 가능)",
        type=["xlsx"], accept_multiple_files=True
    )
with col2:
    ad_files = st.file_uploader(
        "📁 송출요청서 (여러 개 가능)",
        type=["xlsx"], accept_multiple_files=True
    )

if b_files and ad_files:
    if st.button("🚀 실행", type="primary"):
        with st.spinner("분석 중..."):

            ad_file_apts = {}
            for f in ad_files:
                key = ad_key(f.name)
                apts, sheet, col = find_apts_auto(f, apt_map)
                ad_file_apts[key] = apts
                if apts:
                    st.write(f"✅ **{key}**: '{sheet}' 시트 {col+1}번 열 → {len(apts)}개 매칭")
                else:
                    st.write(f"⚠️ **{key}**: 아파트 목록 없음 (패키지 상품)")

            apt_freq = Counter()
            for apts in ad_file_apts.values():
                for a in set(apts):
                    apt_freq[a] += 1

            st.divider()

            for b_file in b_files:
                st.subheader(f"📄 {b_file.name}")
                wb, results = process_b_file(b_file, ad_file_apts, apt_map, apt_freq)

                df_result = pd.DataFrame(results)
                if not df_result.empty:
                    # 유사도 매칭된 행 강조 표시
                    fuzzy_matched = df_result[
                        df_result.get('유사도', pd.Series()).apply(
                            lambda x: isinstance(x, str) and '%' in str(x)
                        )
                    ]
                    if not fuzzy_matched.empty:
                        st.warning(f"⚠️ 유사도 매칭 {len(fuzzy_matched)}건 — 아래 표에서 '매칭 파일' 열로 확인하세요.")

                    st.dataframe(df_result, use_container_width=True)
                    n_ok = len(df_result[df_result['아파트명'] != '—'])
                    n_no = len(df_result) - n_ok
                    c1, c2 = st.columns(2)
                    c1.metric("배정 완료", f"{n_ok}행")
                    c2.metric("미배정", f"{n_no}행")

                out = BytesIO()
                wb.save(out)
                결과파일명 = b_file.name.replace('.xlsx', '_결과.xlsx')
                st.download_button(
                    label=f"📥 {결과파일명} 다운로드",
                    data=out.getvalue(),
                    file_name=결과파일명,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=b_file.name
                )
                st.divider()
else:
    st.info("게첨리스트와 송출요청서를 업로드해 주세요.")
